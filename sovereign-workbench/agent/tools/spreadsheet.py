"""
Spreadsheet Generator Tool for Sovereign AI Workbench.
Reads and writes .xlsx Excel files using openpyxl.
Outputs directly into workspace/deliverables/ folder.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Any, List

WORKSPACE_DIR = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "workspace")
)
DELIVERABLES_DIR = os.path.join(WORKSPACE_DIR, "deliverables")


def create_excel_report(
    sheet_name: str,
    headers: List[str],
    rows: List[List[Any]],
    filename: str = "Inspection_Report.xlsx"
) -> Dict[str, Any]:
    """
    Creates a styled Excel spreadsheet deliverable.
    Returns file details and status.
    """
    try:
        os.makedirs(DELIVERABLES_DIR, exist_ok=True)
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        target_path = os.path.join(DELIVERABLES_DIR, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel max sheet title length is 31

        # Styling definitions
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # Write headers
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row in rows:
            ws.append(row)

        # Style data cells & auto-adjust column width
        for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(target_path)

        if os.path.exists(target_path) and os.path.getsize(target_path) > 0:
            return {
                "status": "success",
                "file_name": filename,
                "file_path": target_path,
                "message": f"Successfully generated {filename} ({os.path.getsize(target_path)} bytes)"
            }
        else:
            return {"status": "error", "error": "Generated Excel file is empty."}

    except Exception as e:
        return {"status": "error", "error": str(e)}


def read_excel_report(file_path: str) -> Dict[str, Any]:
    """Reads content from an Excel file in workspace."""
    try:
        if not os.path.isabs(file_path):
            file_path = os.path.join(WORKSPACE_DIR, file_path)

        if not os.path.exists(file_path):
            return {"status": "error", "error": f"Excel file not found: {file_path}"}

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_data = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(row):
                    rows.append(list(row))
            sheet_data[name] = rows

        return {"status": "success", "sheets": sheet_data}
    except Exception as e:
        return {"status": "error", "error": str(e)}


if __name__ == "__main__":
    res = create_excel_report(
        sheet_name="Unit 4B Summary",
        headers=["Inspection ID", "Component", "Status", "Tested PSI", "Pass/Fail"],
        rows=[
            ["INSP-101", "Valve #1", "Operational", 150, "PASS"],
            ["INSP-102", "Valve #2", "Maintenance Required", 120, "FAIL"],
            ["INSP-103", "Valve #3", "Operational", 150, "PASS"]
        ]
    )
    print(res)
